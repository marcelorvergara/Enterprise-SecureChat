import openpyxl
import xlrd

# Maximum data rows read per sheet. Large operational spreadsheets can have
# tens of thousands of rows; ingesting all of them floods the vector store and
# causes embedding timeouts. 30 rows gives a representative sample of structure
# and values while keeping ingestion fast.
MAX_DATA_ROWS = 30

# Rows grouped per chunk. Batching produces semantically denser chunks that
# score higher in cosine similarity against natural language queries than
# single-row chunks do. 5 rows ≈ 150–300 tokens, well within the 512-token
# chunk limit.
ROWS_PER_CHUNK = 5


def _row_sentence(headers: list[str], cells: list[str]) -> str:
    pairs = [f"{h} is {v}" for h, v in zip(headers, cells) if v]
    return ", ".join(pairs) + "." if pairs else ""


def _make_chunk(headers: list[str], rows: list[list[str]], sheet_name: str) -> dict | None:
    sentences = [_row_sentence(headers, r) for r in rows]
    sentences = [s for s in sentences if s]
    if not sentences:
        return None
    text = f"Sheet: {sheet_name}\n" + "\n".join(sentences)
    return {"text": text, "page_number": None, "sheet_name": sheet_name}


def _parse_xlsx(file_path: str) -> list[dict]:
    """Read .xlsx files using openpyxl (Office Open XML format)."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    chunks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            continue
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(header_row)]
        data_rows: list[list[str]] = []
        for i, row in enumerate(rows_iter):
            if i >= MAX_DATA_ROWS:
                break
            cells = [str(v).strip() if v is not None else "" for v in row]
            if any(cells):
                data_rows.append(cells)
        for i in range(0, len(data_rows), ROWS_PER_CHUNK):
            chunk = _make_chunk(headers, data_rows[i:i + ROWS_PER_CHUNK], sheet_name)
            if chunk:
                chunks.append(chunk)
    wb.close()
    return chunks


def _parse_xls(file_path: str) -> list[dict]:
    """Read .xls files using xlrd (Excel 97-2003 binary format)."""
    wb = xlrd.open_workbook(file_path)
    chunks = []
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        if ws.nrows < 2:
            continue
        headers = [str(ws.cell_value(0, col)).strip() or f"col_{col}"
                   for col in range(ws.ncols)]
        data_rows: list[list[str]] = []
        for row_idx in range(1, min(ws.nrows, MAX_DATA_ROWS + 1)):
            cells = [str(ws.cell_value(row_idx, col)).strip()
                     for col in range(ws.ncols)]
            if any(cells):
                data_rows.append(cells)
        for i in range(0, len(data_rows), ROWS_PER_CHUNK):
            chunk = _make_chunk(headers, data_rows[i:i + ROWS_PER_CHUNK], sheet_name)
            if chunk:
                chunks.append(chunk)
    return chunks


def parse_excel(file_path: str) -> list[dict]:
    """Parse every sheet in a workbook into multi-row natural language chunks.

    Each chunk covers up to ROWS_PER_CHUNK data rows prefixed with the sheet
    name. Natural language format ("Header is Value") produces better semantic
    embeddings than pipe-separated format for retrieval against natural language
    queries. Routes to xlrd for legacy .xls and openpyxl for .xlsx.
    """
    if file_path.lower().endswith(".xls"):
        return _parse_xls(file_path)
    return _parse_xlsx(file_path)
